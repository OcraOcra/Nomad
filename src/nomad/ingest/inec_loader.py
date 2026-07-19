from __future__ import annotations

import logging
import unicodedata
from pathlib import Path
from typing import Any

from nomad.models import Category, HardDataPoint
from nomad.utils import utcnow

logger = logging.getLogger(__name__)


def _norm(text: str) -> str:
    """NFKD + elimina combinantes = comparacion sin acentos."""
    nfkd = unicodedata.normalize("NFKD", str(text))
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _val(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".").replace("\xa0", "").replace(" ", "")
    if not s or s in ("-", "...", "n/a", ""):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


# ---- CBA (Canasta Basica Alimentaria) ----


def parse_cba(path: Path) -> list[HardDataPoint]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    points: list[HardDataPoint] = []
    for sn in wb.sheetnames:
        if "cba" not in sn.lower():
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        data_start = None
        for i, row in enumerate(rows):
            vals = [_safe_str(c) for c in (row or [])]
            if len(vals) >= 6 and vals[1].lower().strip() == "cba" and vals[2]:
                data_start = i
                break
        if data_start is None:
            logger.warning("CBA: no se encontro fila de datos")
            wb.close()
            return points
        for row in rows[data_start:]:
            vals = [_safe_str(c) for c in (row or [])]
            subgrupo = vals[1] if len(vals) > 1 else ""
            if not subgrupo or subgrupo.lower().startswith(("nota", "fuente")):
                continue
            cost_act_total = _val(vals[5]) if len(vals) > 5 else _val(vals[2])
            if cost_act_total is None:
                continue
            slug = _norm(subgrupo).lower().replace(" ", "_")[:40]
            slug = "".join(c for c in slug if c.isalnum() or c == "_").strip("_")
            if not slug:
                continue
            points.append(
                HardDataPoint(
                    name=f"cba_{slug}",
                    value=round(cost_act_total, 0),
                    unit="CRC/mes",
                    period="junio_2026",
                    source="INEC-CBA",
                    url="",
                    category=Category.ECONOMIA,
                    meta={
                        "subgrupo": subgrupo,
                        "cost_ant_total": _val(vals[2]),
                        "cost_ant_urbana": _val(vals[3]),
                        "cost_ant_rural": _val(vals[4]),
                        "cost_act_urbana": _val(vals[6]),
                        "cost_act_rural": _val(vals[7]),
                    },
                )
            )
    wb.close()
    logger.info("CBA: %d puntos", len(points))
    return points


# ---- Pobreza linea de pobreza 2010-2025 ----


def parse_pobreza(path: Path) -> list[HardDataPoint]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    points: list[HardDataPoint] = []
    for sn in wb.sheetnames:
        snl = _norm(sn).lower()
        if "cuadro 1" not in snl and "cuadro 2" not in snl:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows):
            vals = [_safe_str(c) for c in (row or [])]
            label0 = _norm(vals[0]).lower()
            # buscar header exacto: inicia con "region" o "zona y" (fila de titulos, no titulo del cuadro)
            if (label0.startswith("region") or label0.startswith("zona y")) and len(vals[0]) < 40:
                for j in range(i + 1, len(rows)):
                    rvals = [_safe_str(c) for c in (rows[j] or [])]
                    label = rvals[0]
                    if not label or label.startswith("Fuente") or "nota" in label.lower():
                        break
                    if len(rvals) < 3:
                        continue
                    total_pct = _val(rvals[1].replace("\n", ""))
                    if total_pct is None or abs(total_pct - 100) > 0.01:
                        continue
                    pobreza_extrema = _val(rvals[5]) if len(rvals) > 5 else None
                    if pobreza_extrema is not None:
                        slug = _norm(label).lower().replace(" ", "_")[:60]
                        slug = "".join(c for c in slug if c.isalnum() or c == "_").strip("_")
                        meta = {
                            "no_pobres": _val(rvals[2].replace("\n", "")),
                            "pobreza_no_extrema": _val(rvals[4].replace("\n", "")),
                        }
                        points.append(
                            HardDataPoint(
                                name=f"pobreza_extrema_{slug}",
                                value=round(pobreza_extrema, 2),
                                unit="%",
                                period=label,
                                source="INEC-ENAHO",
                                url="",
                                category=Category.ECONOMIA,
                                meta=meta,
                            )
                        )
                    pobreza_total = _val(rvals[3].replace("\n", ""))
                    if pobreza_total is not None and "total pais" in _norm(label).lower():
                        slug = _norm(label).lower().replace(" ", "_")[:60]
                        slug = "".join(c for c in slug if c.isalnum() or c == "_").strip("_")
                        points.append(
                            HardDataPoint(
                                name=f"pobreza_total_{slug}",
                                value=round(pobreza_total, 2),
                                unit="%",
                                period=label,
                                source="INEC-ENAHO",
                                url="",
                                category=Category.ECONOMIA,
                            )
                        )
                break
    wb.close()
    logger.info("Pobreza: %d puntos", len(points))
    return points


# ---- IPM (Pobreza Multidimensional) 2025 ----


def parse_ipm(path: Path) -> list[HardDataPoint]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    points: list[HardDataPoint] = []
    for sn in wb.sheetnames:
        try:
            n = int(sn)
        except ValueError:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        started = False
        for row in rows:
            vals = [_safe_str(c) for c in (row or [])]
            label = vals[0]
            if not label:
                continue
            ll = label.lower().replace("a", "a").replace("o", "o").replace("i", "i")
            if "zona" in ll or "region de plan" in ll:
                started = True
                continue
            if not started:
                continue
            if "fuente:" in label.lower():
                break
            total_abs = _val(vals[1])
            total_rel = _val(vals[2])
            pobre_rel = _val(vals[8]) if len(vals) > 8 else None
            if pobre_rel is not None:
                slug = label.lower().replace(" ", "_")[:60]
                slug = "".join(c for c in slug if c.isalnum() or c == "_").strip("_")
                points.append(
                    HardDataPoint(
                        name=f"ipm_pobre_{slug}",
                        value=round(pobre_rel, 2),
                        unit="%",
                        period="2025",
                        source="INEC-ENAHO-IPM",
                        url="",
                        category=Category.DESARROLLO_CANTONAL,
                        meta={
                            "total_abs": total_abs,
                            "total_rel": total_rel,
                            "pobre_abs": _val(vals[7]),
                            "no_pobre_rel": _val(vals[4]),
                        },
                    )
                )
            if label.lower() == "total de hogares" and pobre_rel is not None:
                points.append(
                    HardDataPoint(
                        name="ipm_hogares_pobres",
                        value=round(pobre_rel, 2),
                        unit="%",
                        period="2025",
                        source="INEC-ENAHO-IPM",
                        url="",
                        category=Category.DESARROLLO_CANTONAL,
                        meta={"total_hogares": total_abs},
                    )
                )
    wb.close()
    logger.info("IPM: %d puntos", len(points))
    return points


# ---- Nacimientos y Defunciones 2025 ----


def parse_nacimientos_defunciones(path: Path) -> list[HardDataPoint]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    points: list[HardDataPoint] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows:
            vals = [_safe_str(c) for c in (row or [])]
            # TOTAL NACIONAL esta en col 2 (vals[2])
            label = vals[2].strip().lower() if len(vals) > 2 else ""
            if "total nacional" in label:
                nac = _val(vals[4]) if len(vals) > 4 else None
                defu = _val(vals[5]) if len(vals) > 5 else None
                mi = _val(vals[6]) if len(vals) > 6 else None
                if nac is not None:
                    points.append(
                        HardDataPoint(
                            name="nacimientos_nacional",
                            value=round(nac),
                            unit="personas",
                            period="2025_preliminar",
                            source="INEC",
                            url="",
                            category=Category.DESARROLLO_CANTONAL,
                            meta={
                                "defunciones": round(defu) if defu else None,
                                "mortalidad_infantil": round(mi) if mi else None,
                            },
                        )
                    )
    wb.close()
    logger.info("Nac/Def: %d puntos", len(points))
    return points


# ---- Empresas (CCSS) Q1 2026 ----


def _parse_ccs_xls(path: Path, prefix: str, unit: str) -> list[HardDataPoint]:
    import xlrd

    wb = xlrd.open_workbook(str(path))
    points: list[HardDataPoint] = []
    for sn in wb.sheet_names():
        ws = wb.sheet_by_name(sn)
        if ws.nrows < 10:
            continue
        for r in range(ws.nrows):
            if r < 6:
                continue
            vals = [ws.cell_value(r, c) for c in range(ws.ncols)]
            label = str(vals[0]).strip() if vals else ""
            if not label or label.startswith("1/") or "fuente:" in label.lower():
                continue
            total = _val(vals[5]) if len(vals) > 5 else _val(vals[1])
            micro = _val(vals[1])
            if total is not None and micro is not None:
                slug = label.lower().replace(" ", "_")[:50]
                slug = "".join(c for c in slug if c.isalnum() or c == "_").strip("_")
                if not slug:
                    continue
                points.append(
                    HardDataPoint(
                        name=f"{prefix}_{slug}",
                        value=total,
                        unit=unit,
                        period="I_trim_2026",
                        source="INEC-CCSS",
                        url="",
                        category=Category.ECONOMIA,
                        meta={
                            "micro": micro,
                            "pequena": _val(vals[2]),
                            "mediana": _val(vals[3]),
                            "grande": _val(vals[4]),
                        },
                    )
                )
    wb.release_resources()
    logger.info("%s: %d puntos", prefix, len(points))
    return points


def parse_empresas(path: Path) -> list[HardDataPoint]:
    return _parse_ccs_xls(path, "empresas", "empresas")


def parse_trabajadores(path: Path) -> list[HardDataPoint]:
    return _parse_ccs_xls(path, "trabajadores", "trabajadores")


# ---- C-14: Incapacidad (CCSS) 2025 ----


def parse_c14_incapacidad(path: Path) -> list[HardDataPoint]:
    import xlrd

    wb = xlrd.open_workbook(str(path))
    points: list[HardDataPoint] = []
    for sn in wb.sheet_names():
        ws = wb.sheet_by_name(sn)
        for r in range(ws.nrows):
            vals = [ws.cell_value(r, c) for c in range(ws.ncols)]
            if r == 9:
                total = _val(vals[1])
                if total is not None:
                    points.append(
                        HardDataPoint(
                            name="incapacidad_dias_promedio_total",
                            value=round(total, 2),
                            unit="dias",
                            period="2025",
                            source="INEC-C14-CCSS",
                            url="",
                            category=Category.SEGURIDAD,
                            meta={
                                "hombres": _val(vals[2]),
                                "mujeres": _val(vals[3]),
                                "total_privada": _val(vals[5]),
                                "total_domestico": _val(vals[9]),
                                "total_autonomo": _val(vals[13]),
                            },
                        )
                    )
    wb.release_resources()
    logger.info("C-14 Incapacidad: %d puntos", len(points))
    return points


# ---- Orquestador ----


# ---- Helpers CSV ----


def _read_csv(path: Path, delimiter: str = ",") -> list[list[str]]:
    import csv

    rows: list[list[str]] = []
    with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
        if delimiter == ";":
            reader = csv.reader(f, delimiter=";")
        else:
            reader = csv.reader(f)
        for row in reader:
            rows.append([c.strip() if c else "" for c in row])
    return rows


def _csv_header_idx(rows: list[list[str]], keywords: list[str]) -> int:
    """Encuentra el indice de la fila de headers por palabras clave."""
    for i, row in enumerate(rows):
        text = " ".join(row).lower()
        if all(k.lower() in text for k in keywords):
            return i
    return 0


def _col(header: list[str], cols: dict[int, str]) -> list[int]:
    """Busca indices de columnas por nombre parcial."""
    result: list[int] = []
    for idx, wanted in cols.items():
        for i, h in enumerate(header):
            if wanted.lower() in h.lower():
                result.append(i)
                break
        else:
            result.append(-1)
    return result


# ---- IPC (Indice de Precios al Consumidor) ----


def parse_ipc(path: Path) -> list[HardDataPoint]:
    rows = _read_csv(path, delimiter=";")
    points: list[HardDataPoint] = []
    hdr_idx = _csv_header_idx(rows, ["fecha", "ipc", "nivel"])
    if hdr_idx >= len(rows):
        return points
    header = rows[hdr_idx]
    cols = _col(header, {0: "fecha", 1: "nivel", 2: "mensual", 3: "interanual", 4: "acumulada"})
    col_date, col_nivel, col_mensual, col_interanual, col_acum = cols
    if col_date < 0:
        return points
    # tomar ultimos 12 meses con variacion interanual
    recent = []
    for row in rows[hdr_idx + 1:]:
        if len(row) <= max(col for col in cols if col >= 0):
            continue
        if col_interanual >= 0 and not row[col_interanual]:
            continue
        recent.append(row)
    last12 = recent[-12:]
    for row in last12:
        fecha = row[col_date] if col_date < len(row) else ""
        nivel = _val(row[col_nivel]) if col_nivel >= 0 else None
        mensual = _val(row[col_mensual]) if col_mensual >= 0 else None
        interanual = _val(row[col_interanual]) if col_interanual >= 0 else None
        acum = _val(row[col_acum]) if col_acum >= 0 else None
        if interanual is not None:
            points.append(
                HardDataPoint(
                    name=f"ipc_variacion_interanual",
                    value=round(interanual, 2),
                    unit="%",
                    period=fecha[:7],
                    source="INEC-IPC",
                    url="",
                    category=Category.ECONOMIA,
                    meta={"nivel": nivel, "mensual": mensual, "acumulada": acum},
                )
            )
    if last12:
        ultima = last12[-1]
        nivel_last = _val(ultima[col_nivel]) if col_nivel >= 0 else None
        if nivel_last:
            points.append(
                HardDataPoint(
                    name="ipc_nivel_actual",
                    value=round(nivel_last, 2),
                    unit="base_dic2020=100",
                    period=ultima[col_date][:7],
                    source="INEC-IPC",
                    url="",
                    category=Category.ECONOMIA,
                )
            )
    logger.info("IPC: %d puntos", len(points))
    return points


# ---- IDS Cantonal ----


def parse_ids_cantonal(path: Path) -> list[HardDataPoint]:
    rows = _read_csv(path)
    points: list[HardDataPoint] = []
    # buscar header con canton e ids
    header = rows[0] if rows else []
    cols = _col(header, {0: "canton", 1: "ids_2023", 2: "provincia"})
    col_canton, col_ids, col_prov = cols
    if col_canton < 0 or col_ids < 0:
        return points
    for row in rows[1:]:
        if len(row) <= max(c for c in cols if c >= 0):
            continue
        canton = row[col_canton] if col_canton < len(row) else ""
        ids_val = _val(row[col_ids])
        prov = row[col_prov] if col_prov < len(row) else ""
        if ids_val is not None and canton:
            points.append(
                HardDataPoint(
                    name=f"ids_{prov.lower().replace(' ','_')}_{canton.lower().replace(' ','_')}"[:70],
                    value=round(ids_val, 2),
                    unit="puntos",
                    period="2023",
                    source="INEC-IDS",
                    url="",
                    category=Category.DESARROLLO_CANTONAL,
                    meta={"provincia": prov, "canton": canton},
                )
            )
    logger.info("IDS: %d puntos", len(points))
    return points


# ---- IDS Dimensiones ----


def parse_ids_dimensiones(path: Path) -> list[HardDataPoint]:
    rows = _read_csv(path)
    points: list[HardDataPoint] = []
    header = rows[0] if rows else []
    cols = _col(header, {0: "canton", 1: "salud", 2: "participa", 3: "seguridad",
                         4: "educacion", 5: "economico", 6: "ids_2023_final", 7: "provincia"})
    col_canton, col_salud, col_part, col_seg, col_edu, col_eco, col_ids, col_prov = cols
    if col_canton < 0:
        return points
    dims = [
        ("salud", col_salud),
        ("participacion", col_part),
        ("seguridad", col_seg),
        ("educacion", col_edu),
        ("economico", col_eco),
    ]
    for row in rows[1:]:
        if len(row) <= max(c for c in cols if c >= 0):
            continue
        canton = row[col_canton] if col_canton < len(row) else ""
        prov = row[col_prov] if col_prov < len(row) else ""
        if not canton:
            continue
        for dim_name, col_d in dims:
            val = _val(row[col_d]) if col_d >= 0 and col_d < len(row) else None
            if val is not None:
                points.append(
                    HardDataPoint(
                        name=f"ids_dim_{dim_name}_{prov.lower()}_{canton.lower()}"[:70],
                        value=round(val, 2),
                        unit="pts",
                        period="2023",
                        source="INEC-IDS",
                        url="",
                        category=Category.DESARROLLO_CANTONAL,
                        meta={"canton": canton, "provincia": prov},
                    )
                )
    logger.info("IDS Dimensiones: %d puntos", len(points))
    return points


# ---- Indicadores Cantonales ArcGIS ----


def parse_indicadores_cantonales(path: Path) -> list[HardDataPoint]:
    rows = _read_csv(path)
    points: list[HardDataPoint] = []
    header = rows[0] if rows else []
    cols = _col(header, {0: "nom_prov", 1: "nom_cant", 2: "idphc", 3: "idhc",
                         4: "idrgc", 5: "igfm", 6: "poblacion", 7: "electores"})
    col_prov, col_canton, col_idphc, col_idhc, col_idrgc, col_igfm, col_pob, col_elec = cols
    if col_canton < 0:
        return points
    indicators = [
        ("idphc", col_idphc, "indice_pobreza_humana"),
        ("idhc", col_idhc, "indice_desarrollo_humano"),
        ("idrgc", col_idrgc, "indice_desarrollo_relativo_genero"),
        ("igfm", col_igfm, "indice_gestion_financiera_municipal"),
        ("poblacion", col_pob, "poblacion_censada"),
        ("electores", col_elec, "electores"),
    ]
    for row in rows[1:]:
        if len(row) <= max(c for c in cols if c >= 0):
            continue
        canton = row[col_canton] if col_canton < len(row) else ""
        prov = row[col_prov] if col_prov < len(row) else ""
        if not canton:
            continue
        for key, col_d, label in indicators:
            val = _val(row[col_d]) if col_d >= 0 and col_d < len(row) else None
            if val is not None:
                pts_label = key.replace("idphc", "idphc").replace("idhc", "idhc")
                points.append(
                    HardDataPoint(
                        name=f"{pts_label}_{prov.lower()}_{canton.lower()}"[:70],
                        value=round(val, 2) if isinstance(val, float) else val,
                        unit="",
                        period="2011",
                        source="INEC-ArcGIS",
                        url="",
                        category=Category.DESARROLLO_CANTONAL,
                        meta={"provincia": prov, "canton": canton},
                    )
                )
    logger.info("Indicadores Cantonales: %d puntos", len(points))
    return points


# ---- PIB Cantonal ----


def parse_pib_cantonal(path: Path) -> list[HardDataPoint]:
    if path.suffix == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        rows_list: list[list[str]] = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows(values_only=True):
                rows_list.append([str(c) if c is not None else "" for c in row])
            break
        wb.close()
        rows = rows_list
    else:
        rows = _read_csv(path)

    points: list[HardDataPoint] = []
    header = rows[0] if rows else []
    cols = _col(header, {0: "anio", 1: "canton", 2: "pib", 3: "provincia",
                         4: "valor_agregado", 5: "exportaciones", 6: "importaciones"})
    col_anio, col_canton, col_pib, col_prov, col_va, col_exp, col_imp = cols
    if col_canton < 0 or col_pib < 0:
        return points

    latest_year = None
    init_year = 2019  # default fallback
    for row in rows[1:]:
        if len(row) <= max(c for c in cols if c >= 0):
            continue
        anio = int(row[col_anio]) if col_anio >= 0 and row[col_anio].isdigit() else init_year
        if latest_year is None or anio > latest_year:
            latest_year = anio

    for row in rows[1:]:
        if len(row) <= max(c for c in cols if c >= 0):
            continue
        anio_str = row[col_anio] if col_anio >= 0 else ""
        try:
            anio = int(anio_str)
        except (ValueError, TypeError):
            continue
        if latest_year and anio != latest_year:
            continue
        canton = row[col_canton] if col_canton < len(row) else ""
        prov = row[col_prov] if col_prov < len(row) else ""
        pib = _val(row[col_pib])
        va = _val(row[col_va]) if col_va >= 0 else None
        if pib is not None and canton:
            points.append(
                HardDataPoint(
                    name=f"pib_{prov.lower()}_{canton.lower()}"[:60],
                    value=round(pib, 0),
                    unit="millones_CRC",
                    period=str(latest_year or "latest"),
                    source="BCCR-PIB-Cantonal",
                    url="",
                    category=Category.ECONOMIA,
                    meta={"provincia": prov, "canton": canton, "valor_agregado": va},
                )
            )
    logger.info("PIB Cantonal: %d puntos (%s)", len(points), latest_year)
    return points


# ---- Estadisticas OIJ (Seguridad) ----


def parse_estadisticas_oij(path: Path) -> list[HardDataPoint]:
    import csv
    from collections import Counter
    from datetime import datetime, timedelta

    aggregator: Counter[str] = Counter()
    prov_counter: Counter[str] = Counter()
    delito_counter: Counter[str] = Counter()
    fecha_counter: Counter[str] = Counter()
    total = 0
    with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return []
        for row in reader:
            delito = (row.get("Delito") or "").strip()
            provincia = (row.get("Provincia") or "").strip()
            fecha_str = (row.get("Fecha") or "").strip()
            canton = (row.get("Canton") or "").strip()
            if delito:
                delito_counter[delito] += 1
            if provincia:
                prov_counter[provincia] += 1
            if canton:
                aggregator[f"{provincia}:{canton}"] += 1
            if fecha_str:
                try:
                    dt = datetime.strptime(fecha_str, "%Y-%m-%d")
                    fecha_counter[dt.strftime("%Y-%m")] += 1
                except ValueError:
                    pass
            total += 1

    if total == 0:
        logger.info("Estadisticas OIJ: vacio")
        return []

    points: list[HardDataPoint] = []
    # total
    points.append(
        HardDataPoint(
            name="delitos_total_registrados",
            value=total,
            unit="casos",
            period="2025-2026",
            source="OIJ-Estadisticas",
            url="",
            category=Category.SEGURIDAD,
        )
    )
    # top delitos
    for delito, count in delito_counter.most_common(8):
        slug = delito.lower().replace(" ", "_")[:50]
        slug = "".join(c for c in slug if c.isalnum() or c == "_").strip("_")
        points.append(
            HardDataPoint(
                name=f"delitos_{slug}",
                value=count,
                unit="casos",
                period="2025-2026",
                source="OIJ",
                url="",
                category=Category.SEGURIDAD,
            )
        )
    # top provincias
    for prov, count in prov_counter.most_common(7):
        points.append(
            HardDataPoint(
                name=f"delitos_provincia_{prov.lower().replace(' ','_')}",
                value=count,
                unit="casos",
                period="2025-2026",
                source="OIJ",
                url="",
                category=Category.SEGURIDAD,
            )
        )
    # top cantones
    for canton_key, count in aggregator.most_common(10):
        points.append(
            HardDataPoint(
                name=f"delitos_canton_{canton_key.lower().replace(':','_').replace(' ','_')}"[:70],
                value=count,
                unit="casos",
                period="2025-2026",
                source="OIJ",
                url="",
                category=Category.SEGURIDAD,
            )
        )
    # tendencia mensual (ultimos 12 meses)
    sorted_months = sorted(fecha_counter.items())
    if sorted_months:
        for month, count in sorted_months[-12:]:
            points.append(
                HardDataPoint(
                    name=f"delitos_mes_{month}",
                    value=count,
                    unit="casos",
                    period=month,
                    source="OIJ",
                    url="",
                    category=Category.SEGURIDAD,
                )
            )

    logger.info("Estadisticas OIJ: %d puntos (%d registros)", len(points), total)
    return points


# ---- Turismo (Cuenta Satelite) ----


def parse_turismo(path: Path) -> list[HardDataPoint]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    points: list[HardDataPoint] = []
    for sn in wb.sheetnames:
        if "cuadro1" not in sn.lower():
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[6:12]:
            vals = [_safe_str(c) for c in (row or [])]
            if not vals[0] or "Productos de consumo" in vals[0]:
                gasto_total = _val(vals[3]) if len(vals) > 3 else _val(vals[1])
                if gasto_total is not None:
                    points.append(
                        HardDataPoint(
                            name="turismo_gasto_receptor_total",
                            value=round(gasto_total, 0),
                            unit="millones_CRC",
                            period="2021",
                            source="BCCR-CST",
                            url="",
                            category=Category.ECONOMIA,
                        )
                    )
                break
    wb.close()
    logger.info("Turismo: %d puntos", len(points))
    return points


# ---- PARSERS y Loader ----


PARSERS: dict[str, tuple[str, Any]] = {
    # XLSX/XLS (originales)
    "cba": ("reEconomCBA", parse_cba),
    "pobreza": ("sepobrezaenaho", parse_pobreza),
    "ipm": ("reenaho2025-ipm", parse_ipm),
    "nacimientos": ("Nacimientos y Defunciones", parse_nacimientos_defunciones),
    "empresas": ("emp_I_", parse_empresas),
    "trabajadores": ("tra_I_", parse_trabajadores),
    "c14": ("C-14_2025", parse_c14_incapacidad),
    # CSV
    "ipc": ("precios", parse_ipc),
    "turs_cst": ("Turismo", parse_turismo),
    "ids_cant": ("ids_cantonal_2023", parse_ids_cantonal),
    "ids_dim": ("ids_dimensiones", parse_ids_dimensiones),
    "ind_cant": ("indicadores_cantonales", parse_indicadores_cantonales),
    "pib_cant": ("pib_cantonal", parse_pib_cantonal),
    "est_oij": ("Estadisticas", parse_estadisticas_oij),
}


def load_inec_data(inec_dir: Path) -> list[HardDataPoint]:
    """Descubre y parsea todos los archivos INEC en el directorio."""
    if not inec_dir.exists():
        logger.info("Directorio INEC no encontrado: %s", inec_dir)
        return []

    all_points: list[HardDataPoint] = []
    for filepath in sorted(inec_dir.iterdir()):
        if not filepath.is_file() or filepath.suffix not in (".xlsx", ".xls", ".csv"):
            continue
        name = filepath.name
        name_norm = _norm(name).lower()
        matched = False
        for key, (needle, parser) in PARSERS.items():
            if _norm(needle).lower() in name_norm:
                logger.info("Parseando %s (%s)...", name, key)
                try:
                    pts = parser(filepath)
                    all_points.extend(pts)
                except Exception as exc:
                    logger.warning("Error parseando %s: %s", name, exc)
                matched = True
                break
        if not matched:
            logger.debug("Sin parser para %s", name)

    logger.info("Total INEC load: %d puntos", len(all_points))
    return all_points
